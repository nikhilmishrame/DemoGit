import openpyxl
Dict={}
book = openpyxl.load_workbook("C:/Users/MISHNIK/Contacts/Desktop/pyrhonDemo.xlsx")
sheet = book.active
data = sheet.cell(row=1, column=1)
print(data.value)


for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_row):
        # print(sheet.cell(row=i,column=j).value)
        Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)



# sheet.cell(row=4, column=2).value = "New"

# book.save("C:/Users/MISHNIK/Contacts/Desktop/pyrhonDemo.xlsx")